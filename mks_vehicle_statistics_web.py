import os
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path
from io import BytesIO

import pandas as pd
import streamlit as st
import pdfplumber
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage

# ---------------------------------------------------------
# Streamlit App Configuration
# ---------------------------------------------------------
APP_TITLE = "MKS Japan Customs Vehicle Export Statistics"
st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")

MKS_BLUE = "#529cba"
MKS_DARK = "#1f3f4f"
MKS_LIGHT = "#eaf5fa"
MKS_WHITE = "#ffffff"

# Custom CSS for Branding
st.markdown(f"""
<style>
    /* Primary text and header color */
    h1, h2, h3, h4, h5, h6 {{
        color: {MKS_DARK} !important;
    }}
    /* Main background */
    .stApp {{
        background-color: {MKS_LIGHT};
    }}
    /* Sidebar styling */
    [data-testid="stSidebar"] {{
        background-color: {MKS_WHITE};
        box-shadow: 2px 0 5px rgba(0,0,0,0.05);
    }}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------
# Core Constants
# ---------------------------------------------------------
MONTHS = ["Jan", "Feb", "Mar", "Apl", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_LABELS = {
    "Jan": "January", "Feb": "February", "Mar": "March", "Apl": "April",
    "May": "May", "Jun": "June", "Jul": "July", "Aug": "August",
    "Sep": "September", "Oct": "October", "Nov": "November", "Dec": "December",
}
QUARTERS = {
    "Q1": ["Jan", "Feb", "Mar"],
    "Q2": ["Apl", "May", "Jun"],
    "Q3": ["Jul", "Aug", "Sep"],
    "Q4": ["Oct", "Nov", "Dec"],
}
PATTERN_MAP = {
    ("Used", "All Vehicles"): r"^ALL DATA \(USED VEHICLE\)(\d{4})\.pdf$",
    ("Used", "Van Only"): r"^ALL DATA \(USED VEHICLE-VAN\)(\d{4})\.pdf$",
    ("New", "All Vehicles"): r"^ALL DATA \(NEW VEHICLE\)(\d{4})\.pdf$",
    ("New", "Van Only"): r"^ALL DATA \(NEW VEHICLE-VAN\)(\d{4})\.pdf$",
}

STEERING_STARTER = {
    "U.S.A.": "Left", "AUSTRALIA": "Right", "CANADA": "Left", "SAUDI ARABIA": "Left", "CHINA": "Left",
    "U.A.E.": "Left", "MEXICO": "Left", "U.K.": "Right", "TAIWAN": "Left", "GERMANY": "Left", "BELGIUM": "Left",
    "POLAND": "Right", "KUWAIT": "Left", "SPAIN": "Left", "NEW ZEALAND": "Right", "PHILIPPINES": "Right",
    "ITALY": "Left", "FRANCE": "Left", "ISRAEL": "Right", "OMAN": "Left", "PUERTO RICO": "Left", "QATAR": "Left",
    "TURKEY": "Left", "CHILE": "Left", "Republic of KOREA": "Left", "RUSSIA": "Left", "TANZANIA": "Right",
    "KENYA": "Right", "MONGOLIA": "Left", "South AFRICA": "Right", "SRI LANKA": "Right", "THAILAND": "Right",
    "PAKISTAN": "Right", "MALAYSIA": "Right", "UGANDA": "Right", "CYPRUS": "Right", "JAMAICA": "Right",
    "NIGERIA": "Right", "GUYANA": "Left", "GHANA": "Right", "BANGLADESH": "Right", "ZAMBIA": "Right",
    "GEORGIA": "Right", "IRELAND": "Left", "MYANMAR": "Right", "MOZAMBIQUE": "Right", "DOMINICAN Republic": "Right",
    "NAMIBIA": "Left", "CAMBODIA": "Right", "LESOTHO": "Left", "BOTSWANA": "Right", "MALTA": "Left", "ZIMBABWE": "Left",
    "JAPAN": "Right", "SINGAPORE": "Right", "BRUNEI": "Right", "INDONESIA": "Right", "HONG KONG": "Left",
    "MACAU": "Left", "NEPAL": "Left", "INDIA": "Right", "ETHIOPIA": "Right", "JORDAN": "Left", "GUATEMALA": "Right",
    "DJIBOUTI": "Right", "COTE D'IVIRE": "Right", "MAURITIUS": "Left", "HONDURAS": "Right", "PANAMA": "Right",
    "PARAGUAY": "Right", "GABON": "Right", "SEYCHELLES": "Left", "NICARAGUA": "Right", "EGYPT": "Left",
    "HAITI": "Right", "SURINAM": "Left",
}

CSV_NAME = "exports_master_all_years.csv"
STEERING_CSV_NAME = "country_steering_rules.csv"
BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------
# Core Helper Functions
# ---------------------------------------------------------
def is_numeric_token(token: str) -> bool:
    return bool(re.fullmatch(r"[\d,]+", token or ""))

def ensure_steering_lookup(base_dir: Path, countries: list[str]) -> Path:
    lookup_path = base_dir / STEERING_CSV_NAME
    if lookup_path.exists():
        try:
            existing = pd.read_csv(lookup_path)
        except Exception:
            existing = pd.DataFrame(columns=["country", "steering_rule", "notes"])
        for col in ["country", "steering_rule", "notes"]:
            if col not in existing.columns:
                existing[col] = ""
        existing["country"] = existing["country"].astype(str)

        missing_rows = []
        existing_countries = set(existing["country"].tolist())
        for country in sorted(set(countries)):
            if country not in existing_countries:
                missing_rows.append({
                    "country": country,
                    "steering_rule": STEERING_STARTER.get(country, "Unknown"),
                    "notes": "",
                })
        if missing_rows:
            existing = pd.concat([existing, pd.DataFrame(missing_rows)], ignore_index=True)
            existing = existing.sort_values("country").reset_index(drop=True)
            existing.to_csv(lookup_path, index=False, encoding="utf-8-sig")
        return lookup_path

    rows = []
    for country in sorted(set(countries)):
        rows.append(
            {
                "country": country,
                "steering_rule": STEERING_STARTER.get(country, "Unknown"),
                "notes": "",
            }
        )
    pd.DataFrame(rows).to_csv(lookup_path, index=False, encoding="utf-8-sig")
    return lookup_path

def discover_master_pdfs(base_dir: Path) -> list[tuple[Path, str, str, int]]:
    discovered = []
    for file_path in base_dir.glob("*.pdf"):
        filename = file_path.name
        for (vehicle_condition, body_scope), pattern in PATTERN_MAP.items():
            m = re.fullmatch(pattern, filename, flags=re.IGNORECASE)
            if m:
                year = int(m.group(1))
                discovered.append((file_path, vehicle_condition, body_scope, year))
                break
    return sorted(discovered, key=lambda x: (x[3], x[1], x[2], x[0].name))

def parse_master_pdf(pdf_path: Path, vehicle_condition: str, body_scope: str, year: int) -> pd.DataFrame:
    rows = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                if (
                    line.startswith("Export Data")
                    or line.startswith("Country Name")
                    or line.startswith("Year Jan")
                    or "printed by" in line
                    or line.startswith("Page ")
                ):
                    continue

                parts = line.split()
                if len(parts) < 14:
                    continue

                numeric_tokens = parts[-13:]
                if not all(is_numeric_token(tok) for tok in numeric_tokens):
                    continue

                country = " ".join(parts[:-13]).strip()
                year_total = int(numeric_tokens[0].replace(",", ""))
                monthly_vals = [int(x.replace(",", "")) for x in numeric_tokens[1:]]

                record = {
                    "country": country,
                    "year": year,
                    "vehicle_condition": vehicle_condition,
                    "body_scope": body_scope,
                    "year_total": year_total,
                    "source_file": pdf_path.name,
                    "page": page_no,
                }
                for mth, val in zip(MONTHS, monthly_vals):
                    record[mth] = val

                rows.append(record)

    df = pd.DataFrame(rows)
    if not df.empty:
        df["sum_check"] = df[MONTHS].sum(axis=1)
        df["sum_matches_year_total"] = df["sum_check"] == df["year_total"]
    return df

@st.cache_data(show_spinner="Building master dataset. This might take a moment if PDFs changed...")
def build_or_load_master_data(base_dir: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """ Wrapped cached function that returns the df and steering dataframe. """
    base_path = Path(base_dir)
    csv_path = base_path / CSV_NAME
    if csv_path.exists():
        df = pd.read_csv(csv_path)
    else:
        discovered = discover_master_pdfs(base_path)
        if not discovered:
            raise FileNotFoundError(
                "No year-based master PDFs were found.\n\n"
                "Expected examples:\n"
                "ALL DATA (USED VEHICLE)2025.pdf\n"
                "ALL DATA (USED VEHICLE-VAN)2025.pdf\n"
                "ALL DATA (NEW VEHICLE)2025.pdf\n"
                "ALL DATA (NEW VEHICLE-VAN)2025.pdf"
            )

        all_frames = []
        for pdf_path, vehicle_condition, body_scope, year in discovered:
            all_frames.append(parse_master_pdf(pdf_path, vehicle_condition, body_scope, year))

        df = pd.concat(all_frames, ignore_index=True)
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    steering_path = ensure_steering_lookup(base_path, sorted(df["country"].dropna().astype(str).unique().tolist()))
    
    steering_df = load_steering_lookup(steering_path)
    return df, steering_df

def load_steering_lookup(csv_path: Path) -> pd.DataFrame:
    if not csv_path.exists():
        return pd.DataFrame(columns=["country", "steering_rule", "notes"])
    df = pd.read_csv(csv_path)
    for col in ["country", "steering_rule", "notes"]:
        if col not in df.columns:
            df[col] = ""
    return df[["country", "steering_rule", "notes"]]

def get_logo_path(base_dir: Path) -> Path | None:
    candidates = [
        "mks_logo.png", "MKS_logo.png", "mks.png", "logo.png", "MKS Logo.png",
    ]
    for name in candidates:
        p = base_dir / name
        if p.exists():
            return p
    return None

def export_result_to_excel_bytes(result: dict, monthly_df: pd.DataFrame, logo_path: Path | None = None) -> bytes:
    summary_rows = [
        {"Field": "Country", "Value": result["country"]},
        {"Field": "Year", "Value": result["year"]},
        {"Field": "Vehicle Condition", "Value": result["vehicle_condition"]},
        {"Field": "Dataset", "Value": result["body_scope"]},
        {"Field": "Period Type", "Value": result["period_type"]},
        {"Field": "Period Selected", "Value": result["period_selected"]},
        {"Field": "Quantity", "Value": result["quantity"]},
        {"Field": "Steering Rule", "Value": result["steering_rule"]},
        {"Field": "Notes", "Value": result["steering_notes"]},
        {"Field": "Source File", "Value": result["source_file"]},
        {"Field": "Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False, startrow=4)
        monthly_df.to_excel(writer, sheet_name="Monthly Breakdown", index=False, startrow=2)

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws1 = wb["Summary"]
    ws2 = wb["Monthly Breakdown"]

    header_fill = PatternFill("solid", fgColor="529CBA")
    dark_fill = PatternFill("solid", fgColor="1F3F4F")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="B7C9D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1["A1"] = "MKS Inspections LLC"
    ws1["A2"] = "Japan Vehicle Export Statistics"
    ws1["A1"].font = Font(size=16, bold=True, color="1F3F4F")
    ws1["A2"].font = Font(size=13, bold=True, color="529CBA")

    for cell in ws1[5]:
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row in ws1.iter_rows(min_row=6, max_row=5 + len(summary_rows), min_col=1, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].font = bold_font

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 44

    ws2["A1"] = "Monthly Breakdown"
    ws2["A1"].font = Font(size=14, bold=True, color="1F3F4F")

    for cell in ws2[3]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row in ws2.iter_rows(min_row=4, max_row=3 + len(monthly_df), min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 14

    if logo_path and logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            img.width = 140
            img.height = 50
            ws1.add_image(img, "D1")
        except Exception:
            pass

    final_buffer = BytesIO()
    wb.save(final_buffer)
    return final_buffer.getvalue()

def get_steering_info(steering_df, country: str) -> tuple[str, str]:
    if steering_df.empty:
        return "Unknown", ""
    matches = steering_df[steering_df["country"].astype(str) == str(country)]
    if matches.empty:
        return "Unknown", ""
    row = matches.iloc[0]
    steering = str(row.get("steering_rule", "")).strip() or "Unknown"
    notes = str(row.get("notes", "")).strip()
    if notes.lower() == "nan":
        notes = ""
    return steering, notes

# ---------------------------------------------------------
# Streamlit Interface
# ---------------------------------------------------------
def main():
    logo_path = get_logo_path(BASE_DIR)
    
    col1, col2 = st.columns([1, 4])
    with col1:
        if logo_path:
            st.image(str(logo_path), width=150)
    with col2:
        st.markdown(f"<h1 style='margin-bottom: 0px;'>MKS Japan Vehicle Export Statistics</h1>", unsafe_allow_html=True)
        st.markdown(f"<p style='color: {MKS_BLUE}; font-size: 1.2em;'>Multi-year Japan Customs export data | New and Used vehicles</p>", unsafe_allow_html=True)
    
    st.divider()

    try:
        master_df, steering_df = build_or_load_master_data(str(BASE_DIR))
        
        master_df["country"] = master_df["country"].astype(str)
        master_df["vehicle_condition"] = master_df["vehicle_condition"].astype(str)
        master_df["body_scope"] = master_df["body_scope"].astype(str)
        master_df["year"] = master_df["year"].astype(int)

        years = sorted(master_df["year"].dropna().unique().tolist())
        countries = sorted(master_df["country"].dropna().unique().tolist())
    except FileNotFoundError as e:
        st.error(str(e))
        return
    except Exception as e:
        st.error(f"Failed to load data: {e}")
        st.exception(e)
        return

    # Sidebar parameters
    st.sidebar.markdown(f"<h3 style='color: {MKS_DARK};'>Search Filters</h3>", unsafe_allow_html=True)
    
    selected_year = st.sidebar.selectbox("Year", reversed(years))
    selected_country = st.sidebar.selectbox("Country", countries)
    selected_condition = st.sidebar.selectbox("Vehicle Condition", ["Used", "New"])
    selected_scope = st.sidebar.selectbox("Dataset", ["All Vehicles", "Van Only"])
    
    period_type = st.sidebar.selectbox("Period Type", ["Full Year", "Quarter", "Month"])
    
    selected_quarter = "Q1"
    selected_month = "Jan"
    
    if period_type == "Quarter":
        selected_quarter = st.sidebar.selectbox("Quarter", list(QUARTERS.keys()))
    elif period_type == "Month":
        selected_month = st.sidebar.selectbox("Month", MONTHS)

    # Perform the search inherently since Streamlit reacts immediately
    subset = master_df[
        (master_df["year"] == selected_year)
        & (master_df["country"] == selected_country)
        & (master_df["vehicle_condition"] == selected_condition)
        & (master_df["body_scope"] == selected_scope)
    ].copy()

    if subset.empty:
        st.info("No matching statistics were found for that selection.")
    else:
        row = subset.iloc[0]

        if period_type == "Full Year":
            selected_months = MONTHS
            period_selected = "Full Year"
            quantity = int(row["year_total"])
        elif period_type == "Quarter":
            selected_months = QUARTERS[selected_quarter]
            period_selected = selected_quarter
            quantity = int(sum(int(row[m]) for m in selected_months))
        else:
            selected_months = [selected_month]
            period_selected = MONTH_LABELS[selected_month]
            quantity = int(row[selected_month])

        steering_rule, steering_notes = get_steering_info(steering_df, selected_country)

        monthly_df = pd.DataFrame([
            {"Month": m, "Month Name": MONTH_LABELS[m], "Quantity": int(row[m])}
            for m in selected_months
        ])

        # Prepare result dict for export
        result = {
            "country": selected_country,
            "year": selected_year,
            "vehicle_condition": selected_condition,
            "body_scope": selected_scope,
            "period_type": period_type,
            "period_selected": period_selected,
            "quantity": quantity,
            "steering_rule": steering_rule,
            "steering_notes": steering_notes,
            "source_file": row["source_file"],
        }
        
        st.markdown(
            f"<div style='background-color: white; padding: 10px 20px; border-radius: 10px; border: 1px solid #c8dce6; margin-bottom: 20px;'>"
            f"<span style='font-size:18px;'>Showing <strong>{selected_year} / {selected_condition} / {selected_scope} / {period_selected}</strong> for <strong>{selected_country}</strong>: {quantity:,}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )

        # Display Results
        c1, c2 = st.columns([1, 1])
        with c1:
            st.markdown(f"<h3 style='color: {MKS_DARK};'>Summary</h3>", unsafe_allow_html=True)
            st.metric("Total Quantity", f"{quantity:,}")
            
            st.markdown(f"**Country**: {selected_country}  \n**Year**: {selected_year}  \n**Vehicle Condition**: {selected_condition}  \n**Dataset**: {selected_scope}  \n**Period**: {period_selected}  \n---\n**Steering Rule**: {steering_rule}  \n**Steering Notes**: {steering_notes if steering_notes else '-'}  \n**Source**: *{row['source_file']}*")
        
        with c2:
            st.markdown(f"<h3 style='color: {MKS_DARK};'>Monthly Breakdown</h3>", unsafe_allow_html=True)
            st.dataframe(monthly_df, use_container_width=True, hide_index=True)
            
            # Export to Excel Button
            st.markdown("<br>", unsafe_allow_html=True)
            excel_data = export_result_to_excel_bytes(result, monthly_df, logo_path)
            
            default_name = (
                f"Japan_Exports_{selected_year}_"
                f"{selected_country.replace(' ', '_').replace('.', '')}_"
                f"{selected_condition}_{selected_scope.replace(' ', '_')}_"
                f"{period_selected.replace(' ', '_')}.xlsx"
            )
            
            st.download_button(
                label="📥 Export to Excel",
                data=excel_data,
                file_name=default_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the summary and monthly breakdown to Excel."
            )

    # Sidebar Lower Actions
    st.sidebar.divider()
    
    # Initialize session state for login
    if "admin_logged_in" not in st.session_state:
        st.session_state["admin_logged_in"] = False

    if not st.session_state["admin_logged_in"]:
        with st.sidebar.expander("🔐 Admin Login"):
            admin_user = st.text_input("Username", key="admin_user")
            admin_pass = st.text_input("Password", type="password", key="admin_pass")
            if st.button("Login", use_container_width=True):
                # NOTE: For production, you should use Streamlit secrets instead of hardcoding
                # e.g., if admin_pass == st.secrets["admin_password"]:
                if admin_user == "admin" and admin_pass == "MksAdmin2026!":
                    st.session_state["admin_logged_in"] = True
                    st.rerun()
                else:
                    st.error("Invalid credentials.")
    else:
        st.sidebar.success("Logged in as Admin")
        if st.sidebar.button("Logout", use_container_width=True):
            st.session_state["admin_logged_in"] = False
            st.rerun()
            
        with st.sidebar.expander("⚙️ Manage Data"):
            st.write("If you add new PDF files to the directory, click the button below to rebuild the cache.")
            if st.button("Refresh Data Cache"):
                csv_path = BASE_DIR / CSV_NAME
                if csv_path.exists():
                    try:
                        csv_path.unlink()
                    except Exception:
                        pass
                build_or_load_master_data.clear()
                st.success("Cache refreshed! Re-running application to build it again...")
                st.rerun()
                
        with st.sidebar.expander("🌎 Edit Steering Rules"):
            st.write("Changes inside the table are saved automatically to the CSV.")
            edited_steering_df = st.data_editor(steering_df, num_rows="dynamic", use_container_width=True)
            # Streamlit re-runs on edit, so we just compare and save
            if not steering_df.equals(edited_steering_df):
                steering_path = BASE_DIR / STEERING_CSV_NAME
                edited_steering_df.to_csv(steering_path, index=False, encoding="utf-8-sig")
                build_or_load_master_data.clear()
                st.success("Saved steering rules. Re-running application...")
                st.rerun()

if __name__ == "__main__":
    main()
